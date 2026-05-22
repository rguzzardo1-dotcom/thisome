"""Weekly Windows desktop cleanup.

Sweeps loose files on the desktop into ``Desktop\\REX CHECK\\<bucket>`` using a
professional, granular folder structure. Classifies each file by filename
keywords and a short content peek so things land where they belong (a "Work
Instructions" PDF goes to Work Instructions, not "Documents").

Safety:
  * Three protected folders are never moved and never recursed into.
  * Any filename or folder name listed in ``skip.txt`` (next to this script)
    is left alone -- use this for "colored" / manually-flagged items.
  * Shows a Tk confirmation dialog with the full plan before any move.
  * ``--dry-run`` prints the plan and exits without prompting or moving.
"""

from __future__ import annotations

import argparse
import os
import re
import shutil
import sys
from dataclasses import dataclass, field
from pathlib import Path

PROTECTED_FOLDERS = {
    "main page - 03. mce operations",
    "2.0 work instructions",
    "sharepoint wdi final link",
}

SWEEP_ROOT_NAME = "REX CHECK"
SKIP_FILE = Path(__file__).with_name("skip.txt")

IMAGE_EXTS = {".jpg", ".jpeg", ".png", ".gif", ".bmp", ".tif", ".tiff", ".webp", ".heic"}
RAW_IMAGE_EXTS = {".cr2", ".nef", ".arw", ".dng"}
VIDEO_EXTS = {".mp4", ".mov", ".avi", ".mkv", ".webm", ".wmv"}
AUDIO_EXTS = {".mp3", ".wav", ".flac", ".m4a", ".ogg"}
ARCHIVE_EXTS = {".zip", ".rar", ".7z", ".tar", ".gz", ".tgz"}
INSTALLER_EXTS = {".exe", ".msi", ".bat", ".cmd", ".ps1"}
SHORTCUT_EXTS = {".lnk", ".url", ".website"}
CODE_EXTS = {".py", ".js", ".ts", ".json", ".yaml", ".yml", ".html", ".css",
             ".ipynb", ".sh", ".sql", ".xml"}
EMAIL_EXTS = {".msg", ".eml"}
SPREADSHEET_EXTS = {".xls", ".xlsx", ".xlsm", ".csv", ".tsv"}
PRESENTATION_EXTS = {".ppt", ".pptx", ".odp"}
DOCUMENT_EXTS = {".doc", ".docx", ".odt", ".rtf", ".txt", ".md"}
PDF_EXTS = {".pdf"}

WORK_INSTRUCTION_KEYS = [
    "sop", "standard operating", "work instruction", "work-instruction",
    "wi-", "procedure", "manual", "how to", "how-to", "training", "checklist",
    "guideline", "policy", "process flow",
]
REPORT_KEYS = [
    "report", "summary", "kpi", "metrics", "dashboard", "weekly", "monthly",
    "quarterly", "ytd", "rollup", "scorecard",
]
INVOICE_KEYS = ["invoice", "receipt", "po-", "purchase order", "quote", "estimate"]
SCREENSHOT_KEYS = ["screenshot", "screen shot", "snip", "capture"]


@dataclass
class Move:
    src: Path
    dst: Path
    bucket: str
    reason: str = ""


@dataclass
class Plan:
    moves: list[Move] = field(default_factory=list)
    skipped: list[tuple[Path, str]] = field(default_factory=list)

    def by_bucket(self) -> dict[str, list[Move]]:
        out: dict[str, list[Move]] = {}
        for m in self.moves:
            out.setdefault(m.bucket, []).append(m)
        return out


def find_desktop() -> Path:
    home = Path.home()
    candidates = [home / "Desktop", home / "OneDrive" / "Desktop"]
    # OneDrive - Company variants
    for child in home.glob("OneDrive*"):
        candidates.append(child / "Desktop")
    for c in candidates:
        if c.is_dir():
            return c
    sys.exit(f"Could not find Desktop under {home}")


def load_skip_list() -> set[str]:
    if not SKIP_FILE.exists():
        return set()
    return {
        line.strip().lower()
        for line in SKIP_FILE.read_text(encoding="utf-8").splitlines()
        if line.strip() and not line.strip().startswith("#")
    }


def is_protected(path: Path, skip_set: set[str]) -> bool:
    name = path.name.lower()
    if name in PROTECTED_FOLDERS:
        return True
    if name in skip_set:
        return True
    return False


def peek_text(path: Path, limit: int = 4096) -> str:
    """Return a lowercase text snippet from the file, or empty string."""
    ext = path.suffix.lower()
    try:
        if ext in {".txt", ".md", ".csv", ".tsv", ".log", ".json", ".xml", ".html"}:
            with path.open("r", encoding="utf-8", errors="ignore") as f:
                return f.read(limit).lower()
        if ext == ".pdf":
            try:
                from pypdf import PdfReader  # optional
                reader = PdfReader(str(path))
                if reader.pages:
                    return (reader.pages[0].extract_text() or "")[:limit].lower()
            except Exception:
                return ""
        if ext == ".docx":
            try:
                from docx import Document  # optional (python-docx)
                doc = Document(str(path))
                return " ".join(p.text for p in doc.paragraphs[:20])[:limit].lower()
            except Exception:
                return ""
        if ext == ".xlsx":
            try:
                from openpyxl import load_workbook  # optional
                wb = load_workbook(str(path), read_only=True, data_only=True)
                sheet = wb.active
                cells: list[str] = []
                for row in sheet.iter_rows(max_row=5, values_only=True):
                    cells.extend(str(c) for c in row if c is not None)
                return " ".join(cells)[:limit].lower()
            except Exception:
                return ""
    except Exception:
        return ""
    return ""


def keyword_hit(haystack: str, keys: list[str]) -> bool:
    return any(k in haystack for k in keys)


def classify(path: Path) -> tuple[str, str]:
    """Return (bucket subpath, reason). Bucket may include subfolders like
    'Pictures/Screenshots'."""
    name = path.name.lower()
    ext = path.suffix.lower()
    stem = path.stem.lower()

    if ext in SHORTCUT_EXTS:
        return "Shortcuts", f"extension {ext}"
    if ext in EMAIL_EXTS:
        return "Saved Emails", f"extension {ext}"
    if ext in INSTALLER_EXTS:
        return "Installers", f"extension {ext}"
    if ext in ARCHIVE_EXTS:
        return "Archives", f"extension {ext}"
    if ext in AUDIO_EXTS:
        return "Audio", f"extension {ext}"
    if ext in VIDEO_EXTS:
        return "Videos", f"extension {ext}"
    if ext in CODE_EXTS:
        return "Code and Scripts", f"extension {ext}"

    if ext in IMAGE_EXTS or ext in RAW_IMAGE_EXTS:
        if keyword_hit(stem, SCREENSHOT_KEYS):
            return "Pictures/Screenshots", "screenshot keyword"
        if ext in RAW_IMAGE_EXTS:
            return "Pictures/Raw Photos", f"raw image {ext}"
        return "Pictures/Photos", f"image {ext}"

    # Filename keyword routing first (cheap)
    if keyword_hit(stem, WORK_INSTRUCTION_KEYS):
        return "Work Instructions", "filename keyword"
    if keyword_hit(stem, INVOICE_KEYS):
        return "Invoices and Receipts", "filename keyword"
    if keyword_hit(stem, REPORT_KEYS):
        return "Reports", "filename keyword"

    # Content peek for docs / PDFs / spreadsheets
    if ext in PDF_EXTS | DOCUMENT_EXTS | SPREADSHEET_EXTS | PRESENTATION_EXTS:
        text = peek_text(path)
        if text:
            if keyword_hit(text, WORK_INSTRUCTION_KEYS):
                return "Work Instructions", "content peek"
            if keyword_hit(text, INVOICE_KEYS):
                return "Invoices and Receipts", "content peek"
            if keyword_hit(text, REPORT_KEYS):
                return "Reports", "content peek"

        if ext in SPREADSHEET_EXTS:
            return "Spreadsheets", f"extension {ext}"
        if ext in PRESENTATION_EXTS:
            return "Presentations", f"extension {ext}"
        if ext in PDF_EXTS:
            return "PDFs", "extension .pdf"
        return "Documents", f"extension {ext}"

    return "Misc", "no rule matched"


def safe_destination(dst_dir: Path, src: Path) -> Path:
    dst = dst_dir / src.name
    if not dst.exists():
        return dst
    stem, suffix = src.stem, src.suffix
    i = 2
    while True:
        candidate = dst_dir / f"{stem} ({i}){suffix}"
        if not candidate.exists():
            return candidate
        i += 1


def build_plan(desktop: Path) -> Plan:
    skip_set = load_skip_list()
    sweep_root = desktop / SWEEP_ROOT_NAME
    plan = Plan()

    for entry in desktop.iterdir():
        if entry == sweep_root:
            continue
        if entry.is_dir():
            if is_protected(entry, skip_set):
                plan.skipped.append((entry, "protected folder"))
            else:
                plan.skipped.append((entry, "folder (not moved)"))
            continue
        if entry.is_symlink() or entry.name.startswith("."):
            plan.skipped.append((entry, "hidden or symlink"))
            continue
        if is_protected(entry, skip_set):
            plan.skipped.append((entry, "in skip.txt"))
            continue

        bucket, reason = classify(entry)
        dst_dir = sweep_root / bucket
        dst = safe_destination(dst_dir, entry)
        plan.moves.append(Move(src=entry, dst=dst, bucket=bucket, reason=reason))

    return plan


def render_plan(plan: Plan) -> str:
    lines = [f"Total moves: {len(plan.moves)}", ""]
    for bucket, moves in sorted(plan.by_bucket().items()):
        lines.append(f"[{bucket}]  ({len(moves)})")
        for m in moves:
            lines.append(f"  {m.src.name}    <-  {m.reason}")
        lines.append("")
    if plan.skipped:
        lines.append("Skipped:")
        for p, why in plan.skipped:
            lines.append(f"  {p.name}    ({why})")
    return "\n".join(lines)


def confirm_via_dialog(summary: str) -> bool:
    try:
        import tkinter as tk
        from tkinter import scrolledtext
    except Exception:
        # No GUI -- fall back to stdin prompt
        print(summary)
        return input("\nProceed with these moves? [y/N]: ").strip().lower() == "y"

    result = {"ok": False}
    root = tk.Tk()
    root.title("Desktop Cleanup -- Confirm")
    root.geometry("720x520")

    tk.Label(root, text="Review the plan below. Nothing has been moved yet.",
             anchor="w").pack(fill="x", padx=10, pady=(10, 4))

    text = scrolledtext.ScrolledText(root, wrap="none")
    text.insert("1.0", summary)
    text.configure(state="disabled")
    text.pack(fill="both", expand=True, padx=10)

    btns = tk.Frame(root)
    btns.pack(fill="x", padx=10, pady=10)

    def on_yes():
        result["ok"] = True
        root.destroy()

    def on_no():
        result["ok"] = False
        root.destroy()

    tk.Button(btns, text="Cancel", width=12, command=on_no).pack(side="right")
    tk.Button(btns, text="Move files", width=14, command=on_yes).pack(side="right", padx=8)

    root.mainloop()
    return result["ok"]


def execute(plan: Plan) -> None:
    for m in plan.moves:
        m.dst.parent.mkdir(parents=True, exist_ok=True)
        shutil.move(str(m.src), str(m.dst))


def main():
    parser = argparse.ArgumentParser(description="Weekly desktop cleanup.")
    parser.add_argument("--dry-run", action="store_true",
                        help="Print the plan and exit without moving anything.")
    parser.add_argument("--yes", action="store_true",
                        help="Skip the confirmation dialog (use with care).")
    args = parser.parse_args()

    desktop = find_desktop()
    print(f"Desktop: {desktop}")
    plan = build_plan(desktop)
    summary = render_plan(plan)
    print(summary)

    if not plan.moves:
        print("\nNothing to move.")
        return

    if args.dry_run:
        print("\n[dry-run] No files were moved.")
        return

    if not args.yes and not confirm_via_dialog(summary):
        print("\nCancelled. No files were moved.")
        return

    execute(plan)
    print(f"\nMoved {len(plan.moves)} file(s) into {desktop / SWEEP_ROOT_NAME}.")


if __name__ == "__main__":
    main()
